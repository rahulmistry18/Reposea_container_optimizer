"""
EXCEL EXPORT — RepoSea Report
================================
Generates a fully formatted, multi-sheet Excel workbook from Gold layer data.

Output: data/exports/reposea_report.xlsx

Sheets produced:
  1. 📊 Dashboard    — KPI summary table, styled for executive review
  2. 📦 Containers   — Full gold table, as Excel Table (filterable, sortable)
  3. 🔴 Critical     — Critical-only filtered view
  4. 🟡 Warning      — Warning-only filtered view
  5. 🛣️ By Trade Lane — Pivot by lane with burn totals
  6. 📋 By Lease     — Pivot by lease type
  7. ℹ️  Schema       — Column definitions for Power BI / Tableau mapping

Power BI / Tableau:
  - Get Data → Excel → Select this file → choose "Containers" sheet
  - All columns match the Gold schema exactly
  - Refresh connection to get latest hourly data

Run:  python -m pipeline.export_excel
"""

import json
import logging
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(asctime)s [EXPORT] %(message)s")
log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill, PatternFill, Side
    )
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    log.warning("openpyxl not installed — install via: pip install openpyxl")

GOLD_DIR    = Path(__file__).parent.parent / "data" / "gold"
EXPORT_DIR  = Path(__file__).parent.parent / "data" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# ── Color Palette ─────────────────────────────────────────────────────────────

DARK_NAVY   = "0C1120"
MID_NAVY    = "111827"
AMBER       = "F59E0B"
AMBER_LIGHT = "FEF3C7"
RED         = "EF4444"
RED_LIGHT   = "FEE2E2"
YELLOW      = "F59E0B"
YELLOW_LIGHT= "FEF9C3"
GREEN       = "22C55E"
GREEN_LIGHT = "DCFCE7"
GRAY_HEADER = "1E2D45"
GRAY_ROW    = "F8FAFC"
WHITE       = "FFFFFF"
TEXT_DARK   = "0F172A"
TEXT_MED    = "475569"
TEXT_LIGHT  = "94A3B8"

# ── Style Helpers ─────────────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color=TEXT_DARK, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)

def align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border() -> Border:
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color="CBD5E1") -> Border:
    return Border(bottom=Side(style="thin", color=color))

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ── Sheet 1: Dashboard / KPI Summary ─────────────────────────────────────────

def write_dashboard_sheet(ws, summary: dict, gold_df: pd.DataFrame):
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False

    # Header banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "REPOSEA — Container Intelligence Dashboard"
    c.font  = Font(bold=True, size=16, color=WHITE, name="Calibri")
    c.fill  = fill(DARK_NAVY)
    c.alignment = align("left", "center")
    ws.row_dimensions[1].height = 40

    # Sub-header: run time
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Pipeline run: {summary.get('generated_at','')[:19].replace('T',' ')} UTC  |  Refreshes hourly via GitHub Actions"
    c.font  = Font(size=10, color=TEXT_LIGHT, name="Calibri", italic=True)
    c.fill  = fill(MID_NAVY)
    c.alignment = align("left", "center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 12   # spacer

    # KPI labels row
    kpis = [
        ("TOTAL CONTAINERS",       summary.get("total_containers", 0),   ""),
        ("🔴 CRITICAL",             summary.get("critical_count", 0),     "Past LFD"),
        ("🟡 WARNING",              summary.get("warning_count", 0),      "≤ 2 days buffer"),
        ("✅ SAFE",                 summary.get("safe_count", 0),         "Buffer > 2 days"),
        ("PENALTY ACCRUED",        f"${summary.get('total_penalty_usd', 0):,.0f}", "USD today"),
        ("AVG BUFFER",             f"{summary.get('avg_buffer_days', 0):.1f} days", "Fleet-wide"),
    ]

    kpi_fills = [GRAY_HEADER, RED, YELLOW, GREEN, DARK_NAVY, GRAY_HEADER]
    kpi_text  = [WHITE,       WHITE, TEXT_DARK, TEXT_DARK, WHITE, WHITE]

    col = 1
    for i, (label, value, sub) in enumerate(kpis):
        # Label
        lc = ws.cell(row=4, column=col, value=label)
        lc.font = Font(bold=True, size=9, color=kpi_text[i], name="Calibri")
        lc.fill = fill(kpi_fills[i])
        lc.alignment = align("center", "center")
        ws.row_dimensions[4].height = 18

        # Value
        vc = ws.cell(row=5, column=col, value=value)
        vc.font = Font(bold=True, size=22, color=kpi_text[i], name="Calibri")
        vc.fill = fill(kpi_fills[i])
        vc.alignment = align("center", "center")
        ws.row_dimensions[5].height = 40

        # Sub-label
        sc = ws.cell(row=6, column=col, value=sub)
        sc.font = Font(size=9, color=kpi_text[i], name="Calibri", italic=True)
        sc.fill = fill(kpi_fills[i])
        sc.alignment = align("center", "center")
        ws.row_dimensions[6].height = 18

        ws.column_dimensions[get_column_letter(col)].width = 18
        col += 1

    ws.row_dimensions[7].height = 12   # spacer

    # ── Trade Lane Summary table ──────────────────────────────────────────────
    row = 8
    lane_headers = ["Trade Lane", "Containers", "Critical", "Warning", "Safe", "Penalty (USD)"]
    for ci, h in enumerate(lane_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill      = fill(GRAY_HEADER)
        c.alignment = align("center", "center")
        c.border    = thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    lane_labels = {"TRANS_PAC":"Trans-Pacific","TRANS_ATL":"Trans-Atlantic",
                   "ASIA_EUR":"Asia-Europe","INTRA_ASIA":"Intra-Asia"}
    for ld in summary.get("by_trade_lane", []):
        lane_name = lane_labels.get(ld["trade_lane"], ld["trade_lane"])
        crit_cnt  = gold_df[gold_df.trade_lane == ld["trade_lane"]]["action_status"].eq("Critical").sum()
        warn_cnt  = gold_df[gold_df.trade_lane == ld["trade_lane"]]["action_status"].eq("Warning").sum()
        safe_cnt  = gold_df[gold_df.trade_lane == ld["trade_lane"]]["action_status"].eq("Safe").sum()
        burn      = gold_df[(gold_df.trade_lane == ld["trade_lane"]) & (gold_df.days_to_lfd <= 0)]["burn_rate_usd"].sum()

        row_data = [lane_name, ld["count"], crit_cnt, warn_cnt, safe_cnt, round(burn, 2)]
        row_fills = [WHITE, WHITE, RED_LIGHT, YELLOW_LIGHT, GREEN_LIGHT, WHITE]
        for ci, (val, fc) in enumerate(zip(row_data, row_fills), 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill      = fill(fc)
            c.alignment = align("center", "center")
            c.border    = thin_border()
            if ci == 6:
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1  # spacer

    # ── Priority segmentation ─────────────────────────────────────────────────
    priorities = [
        ("🔴 HIGH",    "One-Way containers past free time",         "Immediate return — per-diem penalty accruing"),
        ("🟡 MEDIUM",  "Master Leases idle at depot",               "Reposition — active daily rent bleed"),
        ("🟢 LOW",     "Long-Term leases sitting idle",             "Monitor — no immediate daily cost"),
    ]
    pri_fills = [RED_LIGHT, YELLOW_LIGHT, GREEN_LIGHT]

    ws.cell(row=row, column=1, value="Action Priority Segmentation").font = Font(bold=True, size=11, name="Calibri")
    row += 1
    for ci, h in enumerate(["Priority", "Lease Type Condition", "Required Action", "Count"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill = fill(GRAY_HEADER)
        c.alignment = align("center", "center")
    row += 1
    counts = [
        gold_df[(gold_df.action_status=="Critical") & (gold_df.lease_type=="One-Way")].shape[0],
        gold_df[gold_df.lease_type=="Master Lease"].shape[0],
        gold_df[gold_df.lease_type=="Long-Term"].shape[0],
    ]
    for (pri, cond, action), cnt, pf in zip(priorities, counts, pri_fills):
        for ci, val in enumerate([pri, cond, action, cnt], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill(pf)
            c.alignment = align("center" if ci != 3 else "left", "center")
            c.border = thin_border()
        row += 1

    for col_letter, w in zip("ABCDEF", [18, 12, 10, 10, 10, 16]):
        ws.column_dimensions[col_letter].width = w

    set_col_widths(ws, {"A": 22, "B": 14, "C": 12, "D": 12, "E": 12, "F": 18})


# ── Sheet 2: Full Containers Table ────────────────────────────────────────────

def write_containers_sheet(ws, gold_df: pd.DataFrame, title="📦 Containers", filter_status=None):
    ws.title = title
    ws.sheet_view.showGridLines = False

    df = gold_df.copy()
    if filter_status:
        df = df[df.action_status == filter_status]

    # Format display columns
    display_cols = [
        "container_id", "lease_type", "trade_lane_label", "vessel_name",
        "current_loc", "origin_locode", "dest_locode",
        "days_to_lfd", "per_diem_rate", "burn_rate_usd", "action_status",
        "priority_score", "agreed_free_days", "pipeline_run_ts",
    ]
    available = [c for c in display_cols if c in df.columns]
    df = df[available].copy()

    # Human-readable column names
    rename = {
        "container_id":    "Container ID",
        "lease_type":      "Lease Type",
        "trade_lane_label":"Trade Lane",
        "vessel_name":     "Vessel",
        "current_loc":     "Current Location",
        "origin_locode":   "Origin (LOCODE)",
        "dest_locode":     "Destination (LOCODE)",
        "days_to_lfd":     "Days to LFD",
        "per_diem_rate":   "Per Diem (USD)",
        "burn_rate_usd":   "Burn Rate / Penalty (USD)",
        "action_status":   "Status",
        "priority_score":  "Priority Score",
        "agreed_free_days":"Free Days",
        "pipeline_run_ts": "Pipeline Run (UTC)",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # Convert timestamps to string
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).replace("NaT", "").replace("nan", "")

    # Write header row
    header_row = 1
    col_widths = {
        "A": 14, "B": 13, "C": 16, "D": 22, "E": 28,
        "F": 14, "G": 14, "H": 12, "I": 14, "J": 22,
        "K": 12, "L": 13, "M": 11, "N": 22,
    }
    set_col_widths(ws, col_widths)
    freeze(ws, "A2")

    for ci, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=header_row, column=ci, value=col_name)
        c.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill      = fill(DARK_NAVY)
        c.alignment = align("center", "center")
        c.border    = thin_border()
    ws.row_dimensions[header_row].height = 24

    # Write data rows
    status_row_fill = {
        "Critical": RED_LIGHT,
        "Warning":  YELLOW_LIGHT,
        "Safe":     WHITE,
    }
    status_col_idx = list(df.columns).index("Status") + 1 if "Status" in df.columns else None
    days_col_idx   = list(df.columns).index("Days to LFD") + 1 if "Days to LFD" in df.columns else None
    burn_col_idx   = list(df.columns).index("Burn Rate / Penalty (USD)") + 1 if "Burn Rate / Penalty (USD)" in df.columns else None
    perd_col_idx   = list(df.columns).index("Per Diem (USD)") + 1 if "Per Diem (USD)" in df.columns else None

    for ri, (_, row_data) in enumerate(df.iterrows(), 2):
        status_val = str(row_data.get("Status", "Safe"))
        row_fill   = status_row_fill.get(status_val, WHITE)
        alt_fill   = "F8FAFC" if ri % 2 == 0 else WHITE

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(size=10, name="Calibri", color=TEXT_DARK)
            c.alignment = align("center", "center")
            c.border    = bottom_border()

            # Status column — colored text
            if status_col_idx and ci == status_col_idx:
                color_map = {"Critical": RED, "Warning": "B45309", "Safe": "15803D"}
                c.font = Font(bold=True, size=10, name="Calibri",
                              color=color_map.get(status_val, TEXT_DARK))
                c.fill = fill(row_fill)
            # Days to LFD — highlight negative
            elif days_col_idx and ci == days_col_idx:
                try:
                    days_val = float(val)
                    c.number_format = "+0.0;-0.0;0.0"
                    if days_val < 0:
                        c.font = Font(bold=True, size=10, name="Calibri", color=RED)
                    elif days_val <= 2:
                        c.font = Font(bold=True, size=10, name="Calibri", color="B45309")
                except (ValueError, TypeError):
                    pass
            # Currency columns
            elif ci in ([burn_col_idx, perd_col_idx]):
                try:
                    c.value  = float(val)
                    c.number_format = '"$"#,##0.00'
                except (ValueError, TypeError):
                    pass
            else:
                c.fill = fill(alt_fill)

        ws.row_dimensions[ri].height = 18

    # Add Excel Table (enables Power BI auto-detection as a named range)
    if len(df) > 0:
        table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        safe_title = title.replace("📦 ", "").replace("🔴 ", "").replace("🟡 ", "").replace(" ", "_")
        tbl = Table(displayName=f"Table_{safe_title}", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)


# ── Sheet: By Trade Lane ──────────────────────────────────────────────────────

def write_lane_sheet(ws, gold_df: pd.DataFrame):
    ws.title = "🛣️ By Trade Lane"
    ws.sheet_view.showGridLines = False

    lane_map = {"TRANS_PAC":"Trans-Pacific","TRANS_ATL":"Trans-Atlantic",
                "ASIA_EUR":"Asia-Europe","INTRA_ASIA":"Intra-Asia"}

    pivot = (
        gold_df.groupby("trade_lane")
        .agg(
            Total=("container_id", "count"),
            Critical=("action_status", lambda x: (x == "Critical").sum()),
            Warning=("action_status",  lambda x: (x == "Warning").sum()),
            Safe=("action_status",     lambda x: (x == "Safe").sum()),
            Total_Penalty_USD=("burn_rate_usd", lambda x: x[gold_df.loc[x.index, "days_to_lfd"] <= 0].sum()),
            Avg_Buffer_Days=("days_to_lfd", "mean"),
            One_Way=("lease_type", lambda x: (x == "One-Way").sum()),
            Master_Lease=("lease_type", lambda x: (x == "Master Lease").sum()),
            Long_Term=("lease_type", lambda x: (x == "Long-Term").sum()),
        )
        .reset_index()
    )
    pivot["Trade Lane"] = pivot["trade_lane"].map(lane_map).fillna(pivot["trade_lane"])
    pivot = pivot.drop(columns=["trade_lane"])
    pivot = pivot[["Trade Lane","Total","Critical","Warning","Safe",
                   "Total_Penalty_USD","Avg_Buffer_Days","One_Way","Master_Lease","Long_Term"]]
    pivot.columns = ["Trade Lane","Total","Critical","Warning","Safe",
                     "Penalty USD","Avg Buffer (days)","One-Way","Master Lease","Long-Term"]
    pivot = pivot.round({"Penalty USD": 2, "Avg Buffer (days)": 1})
    pivot = pivot.sort_values("Penalty USD", ascending=False)

    headers = list(pivot.columns)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill = fill(DARK_NAVY)
        c.alignment = align("center", "center")
        c.border = thin_border()
    ws.row_dimensions[1].height = 22

    for ri, (_, row) in enumerate(pivot.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = align("center", "center")
            c.border = thin_border()
            c.font = Font(size=10, name="Calibri")
            if headers[ci-1] in ("Penalty USD",):
                try: c.number_format = '"$"#,##0.00'
                except: pass
            elif headers[ci-1] == "Critical" and isinstance(val, (int,float)) and val > 0:
                c.font = Font(bold=True, size=10, name="Calibri", color=RED)
                c.fill = fill(RED_LIGHT)
        ws.row_dimensions[ri].height = 20

    for col_letter, w in zip("ABCDEFGHIJ", [18,8,9,9,8,16,16,9,13,11]):
        ws.column_dimensions[col_letter].width = w

    # Add a bar chart: penalty by lane
    if len(pivot) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Penalty USD by Trade Lane"
        chart.y_axis.title = "Penalty (USD)"
        chart.x_axis.title = "Trade Lane"
        chart.style = 10
        chart.height = 12
        chart.width  = 20

        data = Reference(ws, min_col=6, min_row=1, max_row=len(pivot)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(pivot)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{len(pivot)+4}")


# ── Sheet: By Lease Type ──────────────────────────────────────────────────────

def write_lease_sheet(ws, gold_df: pd.DataFrame):
    ws.title = "📋 By Lease Type"
    ws.sheet_view.showGridLines = False

    pivot = (
        gold_df.groupby("lease_type")
        .agg(
            Total=("container_id", "count"),
            Critical=("action_status", lambda x: (x=="Critical").sum()),
            Warning=("action_status",  lambda x: (x=="Warning").sum()),
            Safe=("action_status",     lambda x: (x=="Safe").sum()),
            Total_Penalty=("burn_rate_usd", lambda x: x[gold_df.loc[x.index,"days_to_lfd"] <= 0].sum()),
            Avg_PerDiem=("per_diem_rate", "mean"),
            Avg_FreeDays=("agreed_free_days", "mean"),
        )
        .reset_index()
    )
    pivot.columns = ["Lease Type","Total","Critical","Warning","Safe",
                     "Total Penalty (USD)","Avg Per Diem (USD)","Avg Free Days"]
    pivot = pivot.round({"Total Penalty (USD)":2,"Avg Per Diem (USD)":2,"Avg Free Days":1})

    for ci, h in enumerate(pivot.columns, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill = fill(DARK_NAVY)
        c.alignment = align("center","center")
        c.border = thin_border()
    ws.row_dimensions[1].height = 22

    lease_fills = {"One-Way": "EFF6FF", "Master Lease": "FAF5FF", "Long-Term": "F0FDF4"}
    for ri, (_, row) in enumerate(pivot.iterrows(), 2):
        lf = lease_fills.get(str(row["Lease Type"]), WHITE)
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill(lf)
            c.alignment = align("center","center")
            c.border = thin_border()
            c.font = Font(size=10, name="Calibri")
            if "USD" in str(list(pivot.columns)[ci-1]):
                try: c.number_format = '"$"#,##0.00'
                except: pass
        ws.row_dimensions[ri].height = 22

    for col_letter, w in zip("ABCDEFGH", [14,8,9,9,8,18,18,12]):
        ws.column_dimensions[col_letter].width = w


# ── Sheet: Schema Reference ───────────────────────────────────────────────────

def write_schema_sheet(ws):
    ws.title = "ℹ️ Schema"
    ws.sheet_view.showGridLines = False

    schema = [
        ("container_id",      "VARCHAR(11)",   "4-letter prefix + 7 digits (e.g. MSKU9082341)"),
        ("contract_id",       "VARCHAR",       "ERP contract reference number"),
        ("lease_type",        "ENUM",          "One-Way | Master Lease | Long-Term"),
        ("trade_lane",        "VARCHAR",       "TRANS_PAC | TRANS_ATL | ASIA_EUR | INTRA_ASIA"),
        ("trade_lane_label",  "VARCHAR",       "Human-readable trade lane name"),
        ("vessel_name",       "VARCHAR",       "Carrying vessel (e.g. Maersk Antares)"),
        ("imo_number",        "INTEGER",       "IMO vessel identifier (7 digits)"),
        ("current_loc",       "VARCHAR",       "UN/LOCODE or 'At Sea (Vessel Name)'"),
        ("origin_locode",     "UN/LOCODE",     "Standardized origin (e.g. CNSHA = Shanghai)"),
        ("dest_locode",       "UN/LOCODE",     "Standardized destination (e.g. USLAX = Los Angeles)"),
        ("eta_parsed",        "TIMESTAMP UTC", "Vessel ETA at destination port"),
        ("lfd_parsed",        "TIMESTAMP UTC", "Contract Last Free Day"),
        ("days_to_lfd",       "FLOAT",         "Buffer = ETA − LFD. NEGATIVE means penalty is accruing"),
        ("agreed_free_days",  "INTEGER",       "Free days as per contract (before per-diem kicks in)"),
        ("per_diem_rate",     "DECIMAL(8,2)",  "Daily penalty/rent rate in USD"),
        ("burn_rate_usd",     "DECIMAL(10,2)", "Critical: total accrued penalty | Others: daily exposure"),
        ("action_status",     "ENUM",          "Critical | Warning | Safe"),
        ("priority_score",    "INTEGER",       "Triage rank (higher = more urgent). Feeds sort order"),
        ("pipeline_run_ts",   "TIMESTAMP UTC", "Audit timestamp of the pipeline run that produced this row"),
    ]

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Gold Layer Schema — Power BI / Tableau Column Reference"
    c.font  = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill  = fill(DARK_NAVY)
    c.alignment = align("left","center")
    ws.row_dimensions[1].height = 30

    headers = ["Column Name", "Data Type", "Description"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill = fill(GRAY_HEADER)
        c.alignment = align("center","center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 20

    for ri, (col, dtype, desc) in enumerate(schema, 3):
        row_fill = GRAY_ROW if ri % 2 == 0 else WHITE
        for ci, val in enumerate([col, dtype, desc], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill(row_fill)
            c.alignment = align("left" if ci == 3 else "center", "center")
            c.border = bottom_border()
            c.font = Font(
                size=10, name="Calibri",
                bold=(ci == 1),
                color="1D4ED8" if ci == 1 else TEXT_MED if ci == 2 else TEXT_DARK
            )
        ws.row_dimensions[ri].height = 18

    # Business logic formula note
    note_row = len(schema) + 4
    ws.merge_cells(f"A{note_row}:D{note_row}")
    c = ws.cell(row=note_row, column=1,
                value="Core Formula:  Buffer (days) = ETA_Destination − Contract_Last_Free_Day  |  "
                      "Buffer < 0 → Critical  |  Buffer ≤ 2 → Warning  |  Buffer > 2 → Safe")
    c.font = Font(size=10, name="Calibri", bold=True, color="92400E", italic=True)
    c.fill = fill(AMBER_LIGHT)
    c.alignment = align("left","center")
    ws.row_dimensions[note_row].height = 22

    set_col_widths(ws, {"A": 22, "B": 16, "C": 70})


# ── Main Export Function ──────────────────────────────────────────────────────

def run() -> Path:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")

    log.info("Loading Gold data...")
    gold_path    = GOLD_DIR / "containers.json"
    summary_path = GOLD_DIR / "summary.json"

    if not gold_path.exists():
        raise FileNotFoundError(f"Gold data not found at {gold_path} — run the pipeline first")

    gold_df  = pd.read_json(gold_path)
    summary  = json.loads(summary_path.read_text()) if summary_path.exists() else {}

    log.info(f"Building Excel workbook ({len(gold_df)} rows)...")
    wb = Workbook()
    wb.remove(wb.active)   # Remove default blank sheet

    # Sheet 1: Dashboard KPIs
    write_dashboard_sheet(wb.create_sheet(), summary, gold_df)

    # Sheet 2: Full containers table (Power BI connects here)
    write_containers_sheet(wb.create_sheet(), gold_df, "📦 Containers")

    # Sheet 3: Critical only
    write_containers_sheet(wb.create_sheet(), gold_df, "🔴 Critical", filter_status="Critical")

    # Sheet 4: Warning only
    write_containers_sheet(wb.create_sheet(), gold_df, "🟡 Warning", filter_status="Warning")

    # Sheet 5: By Trade Lane pivot
    write_lane_sheet(wb.create_sheet(), gold_df)

    # Sheet 6: By Lease Type pivot
    write_lease_sheet(wb.create_sheet(), gold_df)

    # Sheet 7: Schema reference
    write_schema_sheet(wb.create_sheet())

    # Save
    out_path = EXPORT_DIR / "reposea_report.xlsx"
    wb.save(out_path)
    log.info(f"Excel workbook saved → {out_path}")
    log.info(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
    log.info(f"  Size:   {out_path.stat().st_size / 1024:.1f} KB")
    return out_path


if __name__ == "__main__":
    run()
